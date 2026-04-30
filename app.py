from flask import Flask, request, send_file, render_template_string, jsonify
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import zipfile, io, os, smtplib, json
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders

app = Flask(__name__)

COLS_MAP = {
    "EVSE NAME":                 "שם העמדה",
    "MEMBER NAME":               "שם הדייר",
    "MEMBER NUMBER":             "טלפון",
    "CONSUMPTION (KWH)":         "צריכה (KWH)",
    "CHARGING DURATION":         "זמן טעינה",
    "STARTED AT":                "התחלה",
    "ENDED AT":                  "סיום",
    "ENERGY PRICE (WITH TAXES)": "עלות חשמל (₪)",
}

HEADER_FILL = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
ALT_FILL    = PatternFill("solid", start_color="D6E4F0", end_color="D6E4F0")
WHITE_FILL  = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
TOTAL_FILL  = PatternFill("solid", start_color="F0F7FF", end_color="F0F7FF")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
DATA_FONT   = Font(name="Arial", size=10)
BOLD_FONT   = Font(name="Arial", bold=True, size=10)
thin        = Side(style="thin", color="B0C4DE")
BORDER      = Border(left=thin, right=thin, top=thin, bottom=thin)

# Gmail settings from environment variables
GMAIL_USER = os.environ.get("GMAIL_USER", "")
GMAIL_PASS = os.environ.get("GMAIL_PASS", "")

HTML = """<!DOCTYPE html>
<html dir="rtl" lang="he">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>מעבד טעינות חשמל</title>
<style>
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body { font-family: Arial, sans-serif; background: #f4f7fb; color: #1a1a2e; min-height: 100vh; padding: 2rem; }
  .container { max-width: 700px; margin: 0 auto; }
  h1 { font-size: 22px; color: #1F4E79; margin-bottom: 6px; }
  .sub { font-size: 13px; color: #666; margin-bottom: 2rem; }
  .card { background: white; border-radius: 16px; padding: 2rem; box-shadow: 0 4px 24px rgba(0,0,0,0.08); margin-bottom: 1.5rem; }
  .card h2 { font-size: 16px; color: #1F4E79; margin-bottom: 1rem; border-bottom: 1px solid #eef2f7; padding-bottom: 8px; }
  label { font-size: 13px; color: #555; display: block; margin-bottom: 6px; font-weight: bold; }
  .row { margin-bottom: 1.25rem; }
  select, input[type=number], input[type=text], input[type=email], input[type=password] {
    width: 100%; padding: 10px 14px; border: 1px solid #dde3ed;
    border-radius: 8px; font-size: 14px; color: #1a1a2e; background: #f9fbff;
  }
  .month-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 12px; }
  .drop-zone { border: 2px dashed #b0c4de; border-radius: 12px; padding: 2rem; text-align: center;
    cursor: pointer; background: #f4f7fb; transition: all 0.2s; margin-bottom: 1.5rem; }
  .drop-zone:hover { border-color: #1F4E79; background: #e8f0fb; }
  .drop-zone .icon { font-size: 36px; margin-bottom: 8px; }
  .drop-zone .main { font-size: 15px; font-weight: bold; color: #1F4E79; }
  .drop-zone .hint { font-size: 12px; color: #888; margin-top: 4px; }
  #file-name { font-size: 13px; color: #3B6D11; margin-top: 8px; font-weight: bold; }
  input[type=file] { display: none; }
  .btn { width: 100%; padding: 14px; background: #1F4E79; color: white; border: none;
    border-radius: 10px; font-size: 16px; font-weight: bold; cursor: pointer; transition: opacity 0.2s; margin-top: 8px; }
  .btn:hover { opacity: 0.88; }
  .btn.green { background: #2E7D32; }
  .btn:disabled { opacity: 0.5; cursor: default; }
  .partners-table { width: 100%; border-collapse: collapse; font-size: 13px; }
  .partners-table th { background: #1F4E79; color: white; padding: 10px; text-align: right; }
  .partners-table td { padding: 8px 10px; border-bottom: 1px solid #eef2f7; }
  .partners-table tr:nth-child(even) td { background: #f4f7fb; }
  .partners-table input[type=email] { padding: 6px 10px; font-size: 13px; }
  .error { background: #fff0f0; border: 1px solid #f5c1c1; color: #a32d2d;
    padding: 12px 16px; border-radius: 8px; font-size: 13px; margin-bottom: 1rem; }
  .success { background: #f0fff4; border: 1px solid #b7ebc0; color: #2E7D32;
    padding: 12px 16px; border-radius: 8px; font-size: 13px; margin-bottom: 1rem; }
  .spinner { display: none; text-align: center; margin-top: 1rem; font-size: 14px; color: #1F4E79; }
  #step2 { display: none; }
  .status-row td:last-child { font-weight: bold; }
  .status-ok { color: #2E7D32; }
  .status-err { color: #a32d2d; }
</style>
</head>
<body>
<div class="container">
  <h1>⚡ מעבד טעינות חשמל</h1>
  <p class="sub">העלי קובץ, הזיני מיילים ושלחי לכל ועדי הבית בלחיצה אחת</p>

  {% if error %}<div class="error">{{ error }}</div>{% endif %}
  {% if success %}<div class="success">{{ success }}</div>{% endif %}

  <!-- שלב 1: העלאת קובץ -->
  <div class="card" id="step1">
    <h2>שלב 1 — בחרי חודש והעלי קובץ</h2>
    <form method="POST" action="/load" enctype="multipart/form-data" id="form1">
      <div class="month-grid">
        <div class="row">
          <label>חודש</label>
          <select name="month">
            {% for m in months %}<option value="{{ m }}" {% if m == selected_month %}selected{% endif %}>{{ m }}</option>{% endfor %}
          </select>
        </div>
        <div class="row">
          <label>שנה</label>
          <input type="number" name="year" value="{{ selected_year or 2026 }}" min="2020" max="2035" />
        </div>
      </div>
      <div class="drop-zone" onclick="document.getElementById('file-input').click()">
        <div class="icon">📂</div>
        <div class="main">גרגרי קובץ Excel לכאן</div>
        <div class="hint">קובץ ה-Export החודשי</div>
        <div id="file-name"></div>
        <input type="file" name="file" id="file-input" accept=".xlsx,.xls" required />
      </div>
      <button type="submit" class="btn">טעיני קובץ וצרי רשימת ועדי בית ←</button>
    </form>
  </div>

  <!-- שלב 2: מיילים ושליחה -->
  {% if partners %}
  <div class="card" id="step2-card">
    <h2>שלב 2 — הזיני מיילים ושלחי</h2>
    <form method="POST" action="/send" id="form2">
      <input type="hidden" name="month" value="{{ selected_month }}" />
      <input type="hidden" name="year" value="{{ selected_year }}" />

      <table class="partners-table">
        <thead>
          <tr><th>ועד בית</th><th>מספר שורות</th><th>מייל לשליחה</th></tr>
        </thead>
        <tbody>
          {% for p in partners %}
          <tr>
            <td>{{ p.name }}</td>
            <td>{{ p.count }}</td>
            <td><input type="email" name="email_{{ loop.index }}" placeholder="example@gmail.com" value="{{ p.email }}" /></td>
            <input type="hidden" name="partner_{{ loop.index }}" value="{{ p.name }}" />
          </tr>
          {% endfor %}
          <input type="hidden" name="partner_count" value="{{ partners|length }}" />
        </tbody>
      </table>

      <div class="row" style="margin-top:1.5rem;">
        <label>גם להוריד ZIP (בנוסף לשליחה)?</label>
        <select name="also_zip">
          <option value="yes">כן, גם להוריד</option>
          <option value="no">לא, רק לשלוח מיילים</option>
        </select>
      </div>

      <button type="submit" class="btn green" id="send-btn">📧 שלחי מיילים לכל ועדי הבית</button>
      <div class="spinner" id="spinner2">שולחת מיילים, אנא המתיני...</div>
    </form>
  </div>
  {% endif %}

  {% if send_results %}
  <div class="card">
    <h2>תוצאות שליחה</h2>
    <table class="partners-table">
      <thead><tr><th>ועד בית</th><th>סטטוס</th></tr></thead>
      <tbody>
        {% for r in send_results %}
        <tr>
          <td>{{ r.name }}</td>
          <td class="{% if r.ok %}status-ok{% else %}status-err{% endif %}">
            {% if r.ok %}✓ נשלח{% else %}✗ שגיאה: {{ r.error }}{% endif %}
          </td>
        </tr>
        {% endfor %}
      </tbody>
    </table>
  </div>
  {% endif %}

</div>

<script>
  const inp = document.getElementById('file-input');
  const lbl = document.getElementById('file-name');
  const drop = document.querySelector('.drop-zone');

  if (inp) {
    inp.addEventListener('change', () => { lbl.textContent = inp.files[0]?.name || ''; });
    drop.addEventListener('dragover', e => { e.preventDefault(); drop.style.borderColor='#1F4E79'; });
    drop.addEventListener('dragleave', () => drop.style.borderColor='');
    drop.addEventListener('drop', e => {
      e.preventDefault(); drop.style.borderColor='';
      const dt = new DataTransfer(); dt.items.add(e.dataTransfer.files[0]);
      inp.files = dt.files; lbl.textContent = inp.files[0]?.name || '';
    });
  }

  const form2 = document.getElementById('form2');
  if (form2) {
    form2.addEventListener('submit', () => {
      document.getElementById('send-btn').disabled = true;
      document.getElementById('spinner2').style.display = 'block';
    });
  }
</script>
</body>
</html>"""

MONTHS = ["ינואר","פברואר","מרץ","אפריל","מאי","יוני","יולי","אוגוסט","ספטמבר","אוקטובר","נובמבר","דצמבר"]

# שמירת מיילים קבועים
EMAILS_FILE = "/tmp/partner_emails.json"

def get_email_credentials():
    if not GMAIL_USER or not GMAIL_PASS:
        raise ValueError("חסרים פרטי SMTP: יש להגדיר GMAIL_USER ו-GMAIL_PASS בסביבה")
    return GMAIL_USER, GMAIL_PASS

def load_emails():
    if os.path.exists(EMAILS_FILE):
        with open(EMAILS_FILE) as f:
            return json.load(f)
    return {}

def save_emails(emails_dict):
    with open(EMAILS_FILE, "w") as f:
        json.dump(emails_dict, f, ensure_ascii=False)

def build_excel_bytes(subset, partner, label):
    wb = Workbook()
    ws = wb.active
    ws.title = "טעינות " + label
    ws.sheet_view.rightToLeft = True
    num_cols = len(COLS_MAP)
    total_kwh    = subset["צריכה (KWH)"].sum()
    total_energy = subset["עלות חשמל (₪)"].sum()

    ws.merge_cells(f"A1:{get_column_letter(num_cols)}1")
    c = ws["A1"]
    c.value = f"דוח טעינות חשמל לרכבים – {partner} – {label}"
    c.font = Font(name="Arial", bold=True, size=14, color="1F4E79")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells(f"A2:{get_column_letter(num_cols)}2")
    c = ws["A2"]
    c.value = f'סה"כ: {len(subset)} טעינות  |  צריכה כוללת: {total_kwh:.2f} KWH  |  עלות חשמל כוללת: ₪{total_energy:.2f}'
    c.font = Font(name="Arial", size=10, italic=True, color="555555")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 6

    for ci, h in enumerate(subset.columns, 1):
        cell = ws.cell(row=4, column=ci, value=h)
        cell.font = HEADER_FONT; cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[4].height = 28

    for ri, (_, row) in enumerate(subset.iterrows(), 5):
        fill = ALT_FILL if ri % 2 == 0 else WHITE_FILL
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = DATA_FONT; cell.fill = fill
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

    tr = len(subset) + 5
    for ci in range(1, num_cols + 1):
        ws.cell(row=tr, column=ci).fill = TOTAL_FILL
        ws.cell(row=tr, column=ci).border = BORDER
    ws.cell(row=tr, column=1).value = 'סה"כ'
    ws.cell(row=tr, column=1).font = BOLD_FONT
    ws.cell(row=tr, column=1).alignment = Alignment(horizontal="center")
    ws.cell(row=tr, column=4).value = round(total_kwh, 2)
    ws.cell(row=tr, column=4).font = BOLD_FONT
    ws.cell(row=tr, column=4).alignment = Alignment(horizontal="center")
    ws.cell(row=tr, column=num_cols).value = round(total_energy, 2)
    ws.cell(row=tr, column=num_cols).font = BOLD_FONT
    ws.cell(row=tr, column=num_cols).alignment = Alignment(horizontal="center")

    for i, w in enumerate([22,22,16,14,14,20,20,18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def send_email(to_email, partner, label, excel_bytes, safe_name):
    msg = MIMEMultipart()
    msg["From"]    = GMAIL_USER
    msg["To"]      = to_email
    msg["Subject"] = f"דוח טעינות חשמל – {partner} – {label}"

    body = f"""שלום,

מצורף דוח טעינות החשמל לרכבים עבור {partner} לחודש {label}.

בברכה,
מערכת ניהול טעינות חשמל
"""
    msg.attach(MIMEText(body, "plain", "utf-8"))

    part = MIMEBase("application", "octet-stream")
    part.set_payload(excel_bytes)
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f'attachment; filename="{safe_name}.xlsx"')
    msg.attach(part)

    get_email_credentials()

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(GMAIL_USER, GMAIL_PASS)
            server.sendmail(GMAIL_USER, to_email, msg.as_string())
    except smtplib.SMTPAuthenticationError:
        raise ValueError("שגיאת התחברות ל-Gmail SMTP: בדקי את שם המשתמש והסיסמה/סיסמת אפליקציה")
    except Exception as e:
        raise ValueError(f"שגיאת SMTP: {e}")

# Cache קובץ בין requests
import tempfile
CACHE = {}

@app.route("/")
def index():
    return render_template_string(HTML, months=MONTHS, partners=None, send_results=None,
                                   selected_month="מרץ", selected_year=2026, error=None, success=None)

@app.route("/load", methods=["POST"])
def load():
    f = request.files.get("file")
    month = request.form.get("month", "")
    year  = request.form.get("year", "2026")

    if not f:
        return render_template_string(HTML, months=MONTHS, partners=None, send_results=None,
                                       selected_month=month, selected_year=year,
                                       error="נא לבחור קובץ", success=None)
    try:
        df = pd.read_excel(f, sheet_name="Export")
    except Exception:
        return render_template_string(HTML, months=MONTHS, partners=None, send_results=None,
                                       selected_month=month, selected_year=year,
                                       error='לא נמצא גיליון "Export" בקובץ', success=None)

    df = df[df["CONSUMPTION (KWH)"] > 0].copy()
    df_clean = df[["PARTNER"] + list(COLS_MAP.keys())].rename(columns=COLS_MAP)

    # שמור בcache
    CACHE["df"] = df_clean
    CACHE["month"] = month
    CACHE["year"] = year

    saved_emails = load_emails()
    partner_names = [p for p in df_clean["PARTNER"].unique() if p and p != "Evolt_test"]
    partners = [{"name": p, "count": len(df_clean[df_clean["PARTNER"]==p]),
                 "email": saved_emails.get(p, "")} for p in partner_names]

    return render_template_string(HTML, months=MONTHS, partners=partners, send_results=None,
                                   selected_month=month, selected_year=year, error=None, success=None)

@app.route("/send", methods=["POST"])
def send():
    month = request.form.get("month", "")
    year  = request.form.get("year", "2026")
    label = f"{month} {year}"
    also_zip = request.form.get("also_zip", "no")
    count = int(request.form.get("partner_count", 0))

    df_clean = CACHE.get("df")
    if df_clean is None:
        return render_template_string(HTML, months=MONTHS, partners=None, send_results=None,
                                       selected_month=month, selected_year=year,
                                       error="פג תוקף הסשן, אנא העלי את הקובץ מחדש", success=None)

    # שמור מיילים לעתיד
    saved_emails = load_emails()
    partners_data = []
    for i in range(1, count + 1):
        name  = request.form.get(f"partner_{i}", "")
        email = request.form.get(f"email_{i}", "").strip()
        if name:
            partners_data.append({"name": name, "email": email})
            if email:
                saved_emails[name] = email
    save_emails(saved_emails)

    if any(item["email"] for item in partners_data) and (not GMAIL_USER or not GMAIL_PASS):
        return render_template_string(HTML, months=MONTHS, partners=None, send_results=None,
                                       selected_month=month, selected_year=year,
                                       error="חסרים פרטי דואר: יש להגדיר GMAIL_USER ו-GMAIL_PASS ב-Render env", success=None)

    results = []
    zip_buf = io.BytesIO() if also_zip == "yes" else None
    zf = zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) if zip_buf else None

    for pd_item in partners_data:
        partner = pd_item["name"]
        email   = pd_item["email"]
        subset  = df_clean[df_clean["PARTNER"] == partner].drop(columns=["PARTNER"])
        safe    = partner.replace('"','').replace('/','_').replace('\\','_').replace('*','').replace('?','').replace('[','').replace(']','')
        excel_bytes = build_excel_bytes(subset, partner, label)

        if zf:
            zf.writestr(f"{label}/{safe}.xlsx", excel_bytes)

        if email:
            try:
                send_email(email, partner, label, excel_bytes, safe)
                results.append({"name": partner, "ok": True, "error": ""})
            except Exception as e:
                results.append({"name": partner, "ok": False, "error": str(e)})
        else:
            results.append({"name": partner, "ok": False, "error": "לא הוזן מייל"})

    if zf:
        zf.close()
        zip_buf.seek(0)
        return send_file(zip_buf, mimetype="application/zip",
                         as_attachment=True, download_name=f"טעינות_{label}.zip")

    return render_template_string(HTML, months=MONTHS, partners=None, send_results=results,
                                   selected_month=month, selected_year=year, error=None,
                                   success=f"הסתיימה שליחת המיילים לחודש {label}")

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
